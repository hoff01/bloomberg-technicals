#!/usr/bin/env python
"""Independently validate a completed train or score-only technical run."""

from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta, timezone
import hashlib
import importlib.util
import json
from pathlib import Path
import sys
from typing import Any
from zoneinfo import ZoneInfo

import polars as pl
from openpyxl import load_workbook
from pypdf import PdfReader


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.model_artifact import load_model_artifact, model_artifact_path  # noqa: E402
from app.technical_config import (  # noqa: E402
    expected_latest_exchange_session,
    load_technical_config,
)
from app.technical_data import DataPaths  # noqa: E402
from app.technical_labels import (  # noqa: E402
    GASOIL_BBL_PER_MT,
    USD_BBL_TO_CPG_DIVISOR,
)


class Validation:
    def __init__(self) -> None:
        self.rows: list[dict[str, Any]] = []

    def check(self, name: str, passed: bool, detail: str) -> None:
        self.rows.append(
            {"check": name, "status": "PASS" if passed else "FAIL", "detail": detail}
        )

    @property
    def passed(self) -> bool:
        return all(row["status"] == "PASS" for row in self.rows)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT)
    parser.add_argument("--mode", choices=("live", "demo"), required=True)
    parser.add_argument("--workflow", choices=("train", "score"), required=True)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    root = args.project_root.resolve()
    config = load_technical_config(root / "config" / "technical_system.toml")
    paths = DataPaths.under(root, dataset=args.mode)
    dist = root / "dist"
    manifest_path = (
        dist / "technical_scoring_manifest.json"
        if args.workflow == "score"
        else paths.run_manifest
    )
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    validation = Validation()
    validation.check(
        "Pandas-free operating environment",
        importlib.util.find_spec("pandas") is None,
        "pandas is absent" if importlib.util.find_spec("pandas") is None else "pandas is installed",
    )

    expected_workflow = "SCORE_CURRENT" if args.workflow == "score" else "TRAIN_AND_SCORE"
    validation.check(
        "manifest identity",
        manifest.get("mode") == args.mode.upper()
        and manifest.get("workflow") == expected_workflow,
        f"mode={manifest.get('mode')} workflow={manifest.get('workflow')}",
    )
    validation.check(
        "stage timing receipt",
        bool(manifest.get("stage_seconds")),
        json.dumps(manifest.get("stage_seconds") or {}, sort_keys=True),
    )
    workbook_path = root / "Technical_Trading_System.xlsx"
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    validation.check(
        "21-sheet workbook topology",
        len(workbook.sheetnames) == 21
        and workbook.sheetnames[0] == "00 Control"
        and workbook.sheetnames[-1] == "20 Indicator Audit",
        f"sheets={len(workbook.sheetnames)} first={workbook.sheetnames[0]} last={workbook.sheetnames[-1]}",
    )
    workbook.close()

    product_pdf_path = root / "output" / "pdf" / "Technical_Product_Report.pdf"
    pdf_pages = 0
    pdf_text = ""
    pdf_error = ""
    try:
        product_pdf = PdfReader(product_pdf_path)
        pdf_pages = len(product_pdf.pages)
        pdf_text = "\n".join((page.extract_text() or "") for page in product_pdf.pages)
    except Exception as exc:
        pdf_error = f"{type(exc).__name__}: {exc}"
    required_pdf_sections = {
        "Technical Product Report",
        "How the dynamic pattern engine works",
        "NY Harbor ULSD",
        "WTI Crude",
        "Brent Crude",
        "ICE Low Sulphur Gasoil",
        "Bloomberg API readiness",
    }
    missing_pdf_sections = sorted(
        section for section in required_pdf_sections if section not in pdf_text
    )
    validation.check(
        "product-split PDF report",
        product_pdf_path.is_file()
        and product_pdf_path.stat().st_size > 50_000
        and pdf_pages >= 8
        and not missing_pdf_sections
        and not pdf_error,
        (
            f"pages={pdf_pages} bytes="
            f"{product_pdf_path.stat().st_size if product_pdf_path.is_file() else 0} "
            f"missing={missing_pdf_sections} error={pdf_error or 'none'}"
        ),
    )

    spread_library = pl.read_csv(dist / "technical_spread_library.csv")
    model_count = sum(item.model_enabled for item in config.spreads)
    validation.check(
        "spread registry",
        spread_library.height == len(config.spreads)
        and spread_library["spread_id"].n_unique() == len(config.spreads)
        and spread_library.filter(pl.col("model_enabled")).height == model_count,
        (
            f"rows={spread_library.height} unique={spread_library['spread_id'].n_unique()} "
            f"model_enabled={spread_library.filter(pl.col('model_enabled')).height}"
        ),
    )

    signals = pl.read_csv(
        dist / "technical_live_signals.csv",
        try_parse_dates=True,
        infer_schema_length=10_000,
    )
    directional = signals.filter(pl.col("status").is_in(["BUY", "SELL"]))
    directional_valid = (
        directional.is_empty()
        or (
            directional["direction_evidence_validated"].all()
            and not directional["demo_mode"].any()
        )
    )
    demo_safe = args.mode != "demo" or directional.is_empty()
    validation.check(
        "signal board and promotion gate",
        signals.height == len(config.spreads)
        and signals["spread_id"].n_unique() == len(config.spreads)
        and directional_valid
        and demo_safe,
        f"rows={signals.height} directional={directional.height}",
    )
    selected_current = signals.filter(pl.col("portfolio_selected"))
    selected_group_duplicates = (
        selected_current.height
        - selected_current["algebra_group"].n_unique()
        if not selected_current.is_empty()
        else 0
    )
    validation.check(
        "current daily portfolio trade budget",
        selected_current.height
        <= config.backtest.maximum_new_trades_per_session
        and selected_group_duplicates == 0,
        (
            f"selected={selected_current.height} "
            f"limit={config.backtest.maximum_new_trades_per_session} "
            f"duplicate_groups={selected_group_duplicates}"
        ),
    )
    trade_csv_path = root / "output" / "csv" / "Technical_Trade_Levels.csv"
    trade_levels = (
        pl.read_csv(trade_csv_path, try_parse_dates=True, infer_schema_length=10_000)
        if trade_csv_path.is_file()
        else pl.DataFrame()
    )
    required_trade_columns = {
        "trade_code",
        "contract_codes",
        "display_current",
        "display_buy_entry",
        "display_sell_entry",
        "display_fair_value",
        "portfolio_action",
        "quote_convention",
        "calculation_unit",
        "maximum_new_trades_per_session",
        "untouched_lockbox_sessions",
        "lockbox_used_for_training",
    }
    validation.check(
        "broker-facing trade level CSV",
        not trade_levels.is_empty()
        and trade_levels.height == len(config.spreads)
        and trade_levels["spread_id"].n_unique() == len(config.spreads)
        and required_trade_columns.issubset(trade_levels.columns)
        and trade_levels["trade_code"].str.len_chars().min() >= 8
        and not trade_levels["lockbox_used_for_training"].any(),
        (
            f"rows={trade_levels.height} "
            f"columns={len(trade_levels.columns)} path={trade_csv_path}"
        ),
    )
    gasoil_rows = signals.filter(pl.col("gasoil_usd_mt").is_not_null())
    gasoil_bbl_error = (
        float(
            gasoil_rows.select(
                (
                    pl.col("gasoil_usd_bbl")
                    - pl.col("gasoil_usd_mt") / GASOIL_BBL_PER_MT
                )
                .abs()
                .max()
            ).item()
            or 0.0
        )
        if not gasoil_rows.is_empty()
        else float("inf")
    )
    gasoil_cpg_error = (
        float(
            gasoil_rows.select(
                (
                    pl.col("gasoil_cpg")
                    - pl.col("gasoil_usd_mt")
                    / GASOIL_BBL_PER_MT
                    / USD_BBL_TO_CPG_DIVISOR
                )
                .abs()
                .max()
            ).item()
            or 0.0
        )
        if not gasoil_rows.is_empty()
        else float("inf")
    )
    hogo_rows = signals.filter(pl.col("hogo_cpg").is_not_null())
    hogo_display_error = (
        float(
            hogo_rows.select(
                (pl.col("display_current") - pl.col("hogo_cpg"))
                .abs()
                .max()
            ).item()
            or 0.0
        )
        if not hogo_rows.is_empty()
        else float("inf")
    )
    validation.check(
        "gasoil and HOGO unit conversions",
        gasoil_bbl_error <= 1e-5
        and gasoil_cpg_error <= 1e-5
        and hogo_display_error <= 1e-5,
        (
            f"gasoil_rows={gasoil_rows.height} hogo_rows={hogo_rows.height} "
            f"usd_bbl_error={gasoil_bbl_error:.3g} "
            f"cpg_error={gasoil_cpg_error:.3g} "
            f"hogo_display_error={hogo_display_error:.3g}"
        ),
    )
    crack_rows = signals.filter(pl.col("family").str.contains("Crack"))
    crack_display_error = (
        float(
            crack_rows.select(
                (pl.col("display_current") - pl.col("current_spread"))
                .abs()
                .max()
            ).item()
            or 0.0
        )
        if not crack_rows.is_empty()
        else float("inf")
    )
    ho_only_rows = signals.filter(pl.col("structure_roots") == "HO")
    ho_level_pairs = (
        ("display_current", "current_spread"),
        ("display_buy_entry", "buy_entry_ceiling"),
        ("display_sell_entry", "sell_entry_floor"),
        ("display_fair_value", "fair_value_target"),
        ("display_long_stop", "long_stop"),
        ("display_short_stop", "short_stop"),
    )
    ho_cpg_error = max(
        (
            float(
                ho_only_rows.select(
                    (
                        pl.col(display_column)
                        - pl.col(raw_column) / USD_BBL_TO_CPG_DIVISOR
                    )
                    .abs()
                    .max()
                ).item()
                or 0.0
            )
            for display_column, raw_column in ho_level_pairs
        ),
        default=float("inf"),
    )
    validation.check(
        "crack USD per barrel and HO cents-per-gallon quote conventions",
        not crack_rows.is_empty()
        and crack_rows["display_unit"].unique().to_list() == ["USD/bbl"]
        and crack_rows["quote_convention"].unique().to_list()
        == ["CRACK_USD_BBL"]
        and crack_display_error <= 1e-9
        and not ho_only_rows.is_empty()
        and ho_only_rows["display_unit"].unique().to_list() == ["cpg"]
        and ho_only_rows["quote_convention"].unique().to_list() == ["HO_CPG"]
        and ho_cpg_error <= 1e-9,
        (
            f"cracks={crack_rows.height} crack_error={crack_display_error:.3g} "
            f"HO_only={ho_only_rows.height} HO_cpg_error={ho_cpg_error:.3g}"
        ),
    )
    structure_summaries = pl.read_csv(
        dist / "technical_structure_summaries.csv",
        try_parse_dates=True,
        infer_schema_length=10_000,
    )
    model_summary = json.loads(
        (dist / "technical_model_summary.json").read_text(encoding="utf-8")
    )
    required_summary_columns = {
        "current_level",
        "buy_entry_ceiling",
        "sell_entry_floor",
        "fair_value_target",
        "long_stop",
        "short_stop",
        "summary_description",
        "recent_30_trades",
        "recent_30_net_pnl_usd",
        "model_oos_trades",
        "model_net_pnl_usd",
        "pattern_state",
        "pattern_strength",
        "pattern_agreement",
        "pattern_components",
        "trade_code",
        "display_unit",
        "display_current",
        "display_buy_entry",
        "display_sell_entry",
        "display_fair_value",
        "portfolio_action",
    }
    validation.check(
        "trade brief coverage",
        structure_summaries.height == len(config.spreads)
        and structure_summaries["spread_id"].n_unique() == len(config.spreads)
        and required_summary_columns.issubset(structure_summaries.columns)
        and structure_summaries["summary_description"].str.len_chars().min() > 80,
        (
            f"rows={structure_summaries.height} "
            f"unique={structure_summaries['spread_id'].n_unique()}"
        ),
    )
    pattern_columns = {
        "pattern_state",
        "pattern_strength",
        "pattern_agreement",
        "pattern_components",
    }
    pattern_surface_present = pattern_columns.issubset(signals.columns)
    pattern_strength_valid = bool(
        pattern_surface_present
        and signals.filter(
            pl.col("pattern_strength").is_not_null()
            & (
                (pl.col("pattern_strength") < 0)
                | (pl.col("pattern_strength") > 1.000001)
            )
        ).is_empty()
    )
    pattern_agreement_valid = bool(
        pattern_surface_present
        and signals.filter(
            (pl.col("pattern_agreement") < 0)
            | (pl.col("pattern_agreement") > 1.000001)
        ).is_empty()
    )
    validation.check(
        "dynamic multi-indicator pattern surface",
        pattern_surface_present
        and signals["pattern_state"].null_count() == 0
        and signals["pattern_components"].str.len_chars().min() >= 10
        and pattern_strength_valid
        and pattern_agreement_valid,
        (
            f"states={signals['pattern_state'].n_unique()} "
            f"strength_valid={pattern_strength_valid} "
            f"agreement_valid={pattern_agreement_valid}"
        ),
    )
    validation.check(
        "model summary identity",
        model_summary.get("model_id") == manifest.get("model_id"),
        f"model_id={model_summary.get('model_id')}",
    )
    if args.workflow == "score":
        stale_count = signals.filter(pl.col("status") == "MODEL STALE").height
        expected_stale = bool(manifest.get("model_stale", False))
        validation.check(
            "frozen model freshness gate",
            (
                expected_stale
                and stale_count
                == signals.filter(pl.col("model_enabled")).height
            )
            or (not expected_stale and stale_count == 0),
            (
                f"age={manifest.get('model_age_sessions')} "
                f"maximum={manifest.get('maximum_model_age_sessions')} "
                f"stale_rows={stale_count}"
            ),
        )

    quality = pl.read_csv(dist / "technical_data_quality.csv")
    blocking = quality.filter(
        pl.col("blocking") & (pl.col("status") == "FAIL")
    ).height
    validation.check(
        "data quality gates",
        blocking == 0,
        f"blocking_failures={blocking}",
    )

    bars = pl.scan_parquet(paths.bars).filter(pl.col("event_type") == "TRADE")
    bar_summary = bars.select(
        pl.len().alias("rows"),
        pl.col("session_date").n_unique().alias("sessions"),
        pl.col("session_date").max().alias("latest_session"),
        pl.col("pulled_at_utc").max().alias("latest_pull_utc"),
        pl.col("bar_slot").min().alias("min_slot"),
        pl.col("bar_slot").max().alias("max_slot"),
    ).collect().row(0, named=True)
    latest_session = bar_summary["latest_session"]
    latest_pull_utc = bar_summary["latest_pull_utc"]
    pull_local = (
        latest_pull_utc.astimezone(ZoneInfo(config.system.timezone))
        if latest_pull_utc is not None
        else None
    )
    partial_live_session = bool(
        args.mode == "live"
        and pull_local is not None
        and latest_session == pull_local.date()
        and pull_local.time() < config.system.session_end
    )
    if args.mode == "live" and pull_local is not None:
        as_of_date = date.fromisoformat(str(manifest["as_of"]))
        latest_by_root = {
            str(row["root"]): row
            for row in bars.group_by("root")
            .agg(
                pl.col("session_date").max().alias("latest_session"),
                pl.col("bar_end_et").max().alias("latest_bar_end"),
            )
            .collect()
            .to_dicts()
        }
        root_freshness: list[str] = []
        stale_roots: list[str] = []
        for root_code in config.roots:
            expected_session = expected_latest_exchange_session(
                as_of_date,
                pull_local,
                root_code,
                session_start=config.system.session_start,
                bar_interval_minutes=config.system.bar_interval_minutes,
                grace_minutes=config.bloomberg.freshness_grace_minutes,
            )
            latest_row = latest_by_root.get(root_code)
            actual_session = (
                latest_row.get("latest_session") if latest_row else None
            )
            root_freshness.append(
                f"{root_code}={actual_session or 'missing'}/{expected_session}"
            )
            if actual_session is None or actual_session < expected_session:
                stale_roots.append(root_code)
                continue
            latest_bar_end = (
                latest_row.get("latest_bar_end") if latest_row else None
            )
            if latest_bar_end is None:
                stale_roots.append(root_code)
            elif expected_session < pull_local.date():
                latest_end_local = latest_bar_end.astimezone(pull_local.tzinfo)
                if (
                    latest_end_local.date() != expected_session
                    or latest_end_local.time() < config.system.session_end
                ):
                    stale_roots.append(root_code)
                    root_freshness[-1] += " prior_close_missing"
            elif expected_session == pull_local.date():
                close_ready = datetime.combine(
                    pull_local.date(),
                    config.system.session_end,
                    tzinfo=pull_local.tzinfo,
                ) + timedelta(minutes=config.bloomberg.freshness_grace_minutes)
                if pull_local < close_ready:
                    lag_minutes = (
                        pull_local
                        - latest_bar_end.astimezone(pull_local.tzinfo)
                    ).total_seconds() / 60.0
                    maximum_lag = (
                        config.system.bar_interval_minutes
                        + config.bloomberg.freshness_grace_minutes
                    )
                    if lag_minutes < 0 or lag_minutes > maximum_lag:
                        stale_roots.append(root_code)
                        root_freshness[-1] += f" lag={lag_minutes:.1f}m"
                elif latest_bar_end.astimezone(pull_local.tzinfo).time() < config.system.session_end:
                    stale_roots.append(root_code)
                    root_freshness[-1] += " close_bar_missing"
        validation.check(
            "per-root Bloomberg session freshness",
            not stale_roots,
            "; ".join(root_freshness),
        )
    closed_bars = (
        bars.filter(pl.col("session_date") < latest_session)
        if partial_live_session
        else bars
    )
    group_rows = (
        closed_bars.group_by(["security", "session_date"])
        .agg(
            pl.len().alias("bars"),
            pl.col("bar_slot").n_unique().alias("unique_slots"),
            pl.col("bar_slot").min().alias("min_slot"),
            pl.col("bar_slot").max().alias("max_slot"),
        )
        .collect()
    )
    groups = {
        "min": int(group_rows["bars"].min() or 0),
        "max": int(group_rows["bars"].max() or 0),
        "complete_share": float(
            group_rows.select(
                (
                    pl.col("bars")
                    == config.system.complete_bars_per_session
                ).mean()
            ).item()
            or 0.0
        ),
        "duplicate_groups": group_rows.filter(
            pl.col("bars") != pl.col("unique_slots")
        ).height,
        "invalid_slot_groups": group_rows.filter(
            (pl.col("min_slot") < 0)
            | (
                pl.col("max_slot")
                >= config.system.complete_bars_per_session
            )
        ).height,
    }
    validation.check(
        f"complete closed {config.system.bar_interval_minutes}-minute sessions",
        groups["complete_share"] >= 0.80
        and groups["duplicate_groups"] == 0
        and groups["invalid_slot_groups"] == 0,
        f"partial_live_session={partial_live_session}; {bar_summary}; groups={groups}",
    )
    if partial_live_session:
        latest_groups = (
            bars.filter(pl.col("session_date") == latest_session)
            .group_by("security")
            .agg(
                pl.len().alias("bars"),
                pl.col("bar_slot").min().alias("min_slot"),
                pl.col("bar_slot").max().alias("max_slot"),
                pl.col("bar_slot").n_unique().alias("unique_slots"),
                pl.col("bar_end_et").max().alias("latest_bar_end"),
            )
            .with_columns(
                (
                    (pl.col("unique_slots") != pl.col("bars"))
                    | (pl.col("min_slot") < 0)
                    | (
                        pl.col("max_slot")
                        >= config.system.complete_bars_per_session
                    )
                ).alias("bad")
            )
            .collect()
        )
        future_bars = (
            bars.filter(
                (pl.col("session_date") == latest_session)
                & (
                    pl.col("bar_end_et").dt.convert_time_zone("UTC")
                    > pl.col("pulled_at_utc")
                )
            )
            .select(pl.len())
            .collect()
            .item()
        )
        latest_bar_end = latest_groups["latest_bar_end"].max()
        freshness_minutes = (
            (
                latest_pull_utc
                - latest_bar_end.astimezone(timezone.utc)
            ).total_seconds()
            / 60.0
            if latest_pull_utc is not None and latest_bar_end is not None
            else float("inf")
        )
        validation.check(
            "current partial session integrity",
            latest_groups.filter(pl.col("bad")).is_empty()
            and future_bars == 0
            and 0 <= freshness_minutes <= 90,
            (
                f"groups={latest_groups.height} "
                f"bad={latest_groups.filter(pl.col('bad')).height} "
                f"future_bars={future_bars} freshness_minutes={freshness_minutes:.1f}"
            ),
        )

    artifact_file = model_artifact_path(root, args.mode)
    artifact = load_model_artifact(
        artifact_file,
        config=config,
        expected_mode=args.mode.upper(),
    )
    artifact_release_files = dict(artifact.get("training_release_files") or {})
    artifact_release_hashes = dict(artifact.get("training_release_hashes") or {})
    manifest_release_hashes = dict(
        manifest.get("training_release_hashes") or {}
    )
    release_checks: list[str] = []
    release_ok = True
    for key, hash_name in (
        ("seasonality", "seasonality_sha256"),
        ("backtest_trades", "backtest_trades_sha256"),
    ):
        relative_path = artifact_release_files.get(key)
        expected_hash = artifact_release_hashes.get(hash_name)
        if not relative_path or not expected_hash:
            release_ok = False
            release_checks.append(f"{key}=missing")
            continue
        release_path = (artifact_file.parent / str(relative_path)).resolve()
        if not release_path.is_relative_to(artifact_file.parent.resolve()):
            release_ok = False
            release_checks.append(f"{key}=unsafe_path")
            continue
        actual_hash = _sha256(release_path) if release_path.is_file() else None
        matched = actual_hash == expected_hash
        release_ok = release_ok and matched
        release_checks.append(
            f"{key}={'PASS' if matched else 'FAIL'}"
        )
    validation.check(
        "artifact manifest and versioned release identity",
        artifact.get("model_id") == manifest.get("model_id")
        == model_summary.get("model_id")
        and artifact_release_hashes == manifest_release_hashes
        and release_ok,
        (
            f"artifact={artifact.get('model_id')} manifest={manifest.get('model_id')} "
            f"summary={model_summary.get('model_id')} files={release_checks}"
        ),
    )
    training_window = artifact["training_window"]
    validation.check(
        "frozen 30-session lockbox",
        int(training_window["lockbox_sessions"])
        == config.backtest.lockbox_sessions
        and training_window["development_end"] < training_window["lockbox_start"],
        json.dumps(training_window, sort_keys=True),
    )

    if args.workflow == "train":
        feature_scan = pl.scan_parquet(dist / "technical_features.parquet")
        latest_feature_session = feature_scan.select(
            pl.col("session_date").max()
        ).collect().item()
        core_ids = {
            item.spread_id
            for item in config.spreads
            if item.core and item.model_enabled
        }
        latest_core_counts = {
            str(row["spread_id"]): int(row["bars"])
            for row in feature_scan.filter(
                (pl.col("session_date") == latest_feature_session)
                & pl.col("spread_id").is_in(core_ids)
            )
            .group_by("spread_id")
            .agg(pl.len().alias("bars"))
            .collect()
            .to_dicts()
        }
        bad_core = sorted(
            spread_id
            for spread_id in core_ids
            if latest_core_counts.get(spread_id)
            != config.system.complete_bars_per_session
        )
        validation.check(
            "complete latest core packages",
            not bad_core,
            f"latest_session={latest_feature_session} incomplete={bad_core}",
        )
        latest_scale_error = (
            pl.scan_parquet(dist / "technical_features.parquet")
            .sort("timestamp_utc")
            .group_by("spread_id", maintain_order=True)
            .tail(1)
            .select(
                (pl.col("research_close") - pl.col("spread_close"))
                .abs()
                .max()
            )
            .collect()
            .item()
        )
        validation.check(
            "current contract scale anchor",
            float(latest_scale_error or 0.0) <= 1e-9,
            f"maximum_latest_scale_error={float(latest_scale_error or 0.0)}",
        )
        scorecard = pl.read_csv(dist / "technical_strategy_scorecard.csv")
        expected_trials = model_count * scorecard["strategy_id"].n_unique()
        validation.check(
            "complete multiple-testing ledger",
            scorecard.height == expected_trials
            and scorecard["trial_id"].n_unique() == expected_trials,
            f"rows={scorecard.height} expected={expected_trials}",
        )
        validation.check(
            "lockbox excluded from strategy selection",
            "lockbox_used_for_selection" in scorecard.columns
            and not scorecard["lockbox_used_for_selection"].any(),
            (
                f"rows={scorecard.height} "
                f"lockbox_selection_rows="
                f"{scorecard.filter(pl.col('lockbox_used_for_selection')).height if 'lockbox_used_for_selection' in scorecard.columns else 'missing'}"
            ),
        )
        windows = pl.read_csv(
            dist / "technical_backtest_windows.csv", try_parse_dates=True
        )
        window_overlap = windows.filter(
            (pl.col("train_end") >= pl.col("lockbox_start"))
            | (pl.col("validation_end") >= pl.col("lockbox_start"))
            | (pl.col("test_end") >= pl.col("lockbox_start"))
            | pl.col("lockbox_used_for_training")
        ).height
        validation.check(
            "larger walk-forward windows and hard 30-session holdout",
            windows.height >= 2
            and windows["lockbox_sessions"].unique().to_list()
            == [config.backtest.lockbox_sessions]
            and windows["available_sessions"].min()
            >= config.system.minimum_production_sessions
            and windows["train_sessions_in_fold"].min()
            == config.backtest.train_sessions
            and windows["validation_sessions_in_fold"].unique().to_list()
            == [config.backtest.validation_sessions]
            and windows["embargo_sessions"].unique().to_list()
            == [config.backtest.embargo_sessions]
            and windows["oos_test_sessions"].unique().to_list()
            == [config.backtest.test_sessions]
            and window_overlap == 0,
            (
                f"folds={windows.height} "
                f"available_sessions={windows['available_sessions'].min()} "
                f"train_sessions={config.backtest.train_sessions} "
                f"validation_sessions={config.backtest.validation_sessions} "
                f"embargo_sessions={config.backtest.embargo_sessions} "
                f"oos_sessions={config.backtest.test_sessions} "
                f"overlaps={window_overlap}"
            ),
        )
        hogo_fly_count = spread_library.filter(
            pl.col("family") == "HOGO Fly"
        ).height
        validation.check(
            "HOGO outright, box, and fly curve coverage",
            hogo_fly_count >= 14
            and spread_library.filter(
                pl.col("spread_id").str.starts_with("RV.HO_QS.")
            ).height
            >= 15
            and spread_library.filter(
                pl.col("spread_id").str.starts_with("BOX.CURVE.HO_QS.")
            ).height
            >= 14,
            f"HOGO_flies={hogo_fly_count}",
        )
        trades = pl.read_csv(
            dist / "technical_backtest_trades.csv",
            try_parse_dates=True,
            infer_schema_length=10_000,
        )
        entry_violations = trades.filter(
            pl.col("entry_time") <= pl.col("entry_signal_time")
        ).height
        expiry_violations = trades.filter(
            pl.col("exit_session") > pl.col("forced_exit_session")
        ).height
        forced = trades.filter(pl.col("exit_reason") == "MANDATORY_D4_EXIT")
        forced_mismatch = forced.filter(
            (pl.col("exit_session") != pl.col("forced_exit_session"))
            | (
                pl.col("exit_time")
                .dt.convert_time_zone(config.system.timezone)
                .dt.hour()
                != 14
            )
            | (
                pl.col("exit_time")
                .dt.convert_time_zone(config.system.timezone)
                .dt.minute()
                != 15
            )
        ).height
        validation.check(
            "next-bar execution and D-4 liquidation",
            entry_violations == 0
            and expiry_violations == 0
            and forced_mismatch == 0,
            (
                f"trades={trades.height} entry_violations={entry_violations} "
                f"expiry_violations={expiry_violations} "
                f"forced_exit_mismatches={forced_mismatch}"
            ),
        )
        portfolio_lockbox = pl.read_csv(
            dist / "technical_portfolio_lockbox_trades.csv",
            try_parse_dates=True,
            infer_schema_length=10_000,
        )
        selected_lockbox = portfolio_lockbox.filter(
            pl.col("portfolio_selected")
        )
        maximum_daily_entries = (
            int(
                selected_lockbox.group_by("entry_session")
                .len()["len"]
                .max()
                or 0
            )
            if not selected_lockbox.is_empty()
            else 0
        )
        duplicate_budget_groups = (
            selected_lockbox.group_by("entry_session", "algebra_group")
            .len()
            .filter(pl.col("len") > 1)
            .height
            if not selected_lockbox.is_empty()
            else 0
        )
        validation.check(
            "portfolio lockbox daily trade budget",
            maximum_daily_entries
            <= config.backtest.maximum_new_trades_per_session
            and duplicate_budget_groups == 0
            and not portfolio_lockbox["lockbox_used_for_selection"].any(),
            (
                f"selected={selected_lockbox.height} "
                f"max_daily={maximum_daily_entries} "
                f"duplicate_groups={duplicate_budget_groups}"
            ),
        )
        adaptive = pl.read_csv(
            dist / "technical_adaptive_weight_history.csv", try_parse_dates=True
        )
        weight_columns = [
            column
            for column in adaptive.columns
            if column.startswith("adaptive_weight_")
        ]
        weight_metrics = adaptive.select(
            (pl.sum_horizontal(*weight_columns) - 1.0)
            .abs()
            .max()
            .alias("max_sum_error"),
            pl.max_horizontal(*weight_columns).max().alias("max_weight"),
            pl.min_horizontal(*weight_columns).min().alias("min_weight"),
        ).row(0, named=True)
        validation.check(
            "adaptive weight constraints",
            weight_metrics["max_sum_error"] <= 1e-5
            and weight_metrics["max_weight"]
            <= config.backtest.adaptive_max_expert_weight + 1e-6
            and weight_metrics["min_weight"] >= 0,
            str(weight_metrics),
        )
    else:
        parity = manifest.get("train_score_parity") or {}
        validation.check(
            "frozen train-score parity",
            parity.get("status")
            in {"PASS", "NOT_APPLICABLE_SOURCE_CHANGED"},
            json.dumps(parity, sort_keys=True),
        )

    for key, source in (
        ("bars_sha256", paths.bars),
        ("daily_sha256", paths.daily),
        ("contracts_sha256", paths.contracts),
    ):
        expected = str(manifest["files"][key])
        actual = _sha256(source)
        validation.check(
            f"manifest hash {key}",
            actual == expected,
            f"expected={expected[:12]} actual={actual[:12]}",
        )

    payload = {
        "status": "PASS" if validation.passed else "FAIL",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "mode": args.mode.upper(),
        "workflow": expected_workflow,
        "checks": validation.rows,
    }
    output = (
        args.output.resolve()
        if args.output is not None
        else dist / "technical_release_validation.json"
    )
    output.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    for row in validation.rows:
        print(f"{row['status']:4s}  {row['check']}: {row['detail']}")
    print(f"Overall: {payload['status']} ({output})")
    return 0 if validation.passed else 1


if __name__ == "__main__":
    raise SystemExit(main())
