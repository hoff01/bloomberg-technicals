#!/usr/bin/env python
"""Build the 21-sheet technical trading workbook from validated Polars outputs."""

from __future__ import annotations

import argparse
from datetime import date, datetime, timezone
import json
import math
import os
from pathlib import Path
import re
import sys
import tempfile
from typing import Iterable

import polars as pl
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parents[1]
NAVY = "173B57"
BLUE = "2F6FB0"
ORANGE = "F28C45"
PALE_BLUE = "EAF1F6"
PALE_GREEN = "DDF3E8"
PALE_RED = "F8DEDC"
PALE_AMBER = "FFF0CC"
GRAY = "637381"
WHITE = "FFFFFF"
LINE = Side(style="thin", color="DCE4EA")

# openpyxl styles are immutable after assignment, so sharing these objects avoids
# rebuilding and re-hashing identical style descriptors for every exported cell.
FRAME_HEADER_FONT = Font(name="Aptos", size=9, bold=True, color=WHITE)
FRAME_HEADER_FILL = PatternFill("solid", fgColor=BLUE)
FRAME_HEADER_ALIGNMENT = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
FRAME_HEADER_BORDER = Border(bottom=LINE)
FRAME_BODY_FONT = Font(name="Aptos", size=9, color="1F2933")
FRAME_BODY_BORDER = Border(bottom=LINE)
FRAME_SUMMARY_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
FRAME_ALTERNATE_FILL = PatternFill("solid", fgColor="F7F9FB")

TRADE_BRIEF_PREFERRED_COLUMNS = (
    "spread_id",
    "trade_code",
    "contract_codes",
    "structure_roots",
    "quote_convention",
    "calculation_unit",
    "portfolio_action",
    "portfolio_selected",
    "current_status",
    "display_unit",
    "display_current",
    "display_buy_entry",
    "display_sell_entry",
    "display_fair_value",
    "display_long_stop",
    "display_short_stop",
    "confidence",
    "current_level",
    "buy_entry_ceiling",
    "sell_entry_floor",
    "fair_value_target",
    "long_stop",
    "short_stop",
    "heating_oil_cpg",
    "gasoil_usd_mt",
    "gasoil_usd_bbl",
    "gasoil_cpg",
    "hogo_cpg",
    "pattern_state",
    "pattern_strength",
    "pattern_agreement",
    "pattern_components",
    "advanced_risk_regime",
    "relationship_health_scope",
    "tod_normalized_change",
    "vol_regime_ratio_1d_20d",
    "liquidity_stress_ratio",
    "tail_event_rate_20d",
    "robust_volume_surprise",
    "return_skew_5d",
    "return_excess_kurtosis_5d",
    "realized_vol_of_vol_5d",
    "trend_hac_t_stat_3d",
    "close_path_choppiness_5d",
    "selected_strategy_name",
    "selected_strategy_status",
    "model_oos_trades",
    "model_net_pnl_usd",
    "recent_30_trades",
    "recent_30_net_pnl_usd",
    "summary_description",
)


def _frame(path: Path) -> pl.DataFrame:
    if not path.is_file() or path.stat().st_size == 0:
        return pl.DataFrame()
    return pl.read_csv(path, try_parse_dates=True, infer_schema_length=10_000)


def _excel_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            return value.astimezone(timezone.utc).replace(tzinfo=None)
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, float) and not math.isfinite(value):
        return None
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


def _table_name(title: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", title)
    if not cleaned or cleaned[0].isdigit():
        cleaned = "T_" + cleaned
    return cleaned[:240]


def _title_block(ws, title: str, subtitle: str, width: int) -> None:
    end = max(7, width)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end)
    ws.cell(1, 1, title)
    ws.cell(2, 1, subtitle)
    ws.cell(1, 1).font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    ws.cell(2, 1).font = Font(name="Aptos", size=10, color=WHITE)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 27
    ws.row_dimensions[2].height = 30


def _configure_print(ws, max_row: int, max_column: int) -> None:
    print_column = max(1, min(16, max_column))
    print_row = max(5, min(54, max_row))
    ws.print_area = f"A1:{get_column_letter(print_column)}{print_row}"
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.oddFooter.center.text = "Bloomberg Technicals - &[Tab] - Page &[Page] of &[Pages]"
    ws.oddFooter.center.size = 8
    ws.oddFooter.center.color = GRAY
    ws.sheet_view.zoomScale = 85


def _number_format(column: str) -> str:
    key = column.lower()
    if (
        key.endswith("_date")
        or key.endswith("_session")
        or key in {"session_date", "delivery_month", "as_of"}
    ):
        return "yyyy-mm-dd"
    if "timestamp" in key or key.endswith("_time") or key.endswith("_utc"):
        return "yyyy-mm-dd hh:mm"
    if any(token in key for token in ("confidence", "share", "rate", "probability", "coverage")):
        return "0.0%"
    if any(token in key for token in ("pnl", "cost", "edge_usd", "drawdown", "expectancy")):
        return '$#,##0.00;[Red]-$#,##0.00'
    if any(token in key for token in ("count", "trades", "sessions", "bars", "observations", "tier", "rank", "structures", "validated")):
        return "#,##0"
    return "0.0000"


def _write_frame_sheet(
    workbook: Workbook,
    name: str,
    title: str,
    subtitle: str,
    frame: pl.DataFrame,
    *,
    limit: int | None = None,
) -> object:
    ws = workbook.create_sheet(name)
    selected = frame.head(limit) if limit is not None else frame
    columns = selected.columns
    _title_block(ws, title, subtitle, len(columns))
    if not columns:
        ws.cell(4, 1, "No rows available")
        return ws
    for index, column in enumerate(columns, start=1):
        cell = ws.cell(4, index, column)
        cell.font = FRAME_HEADER_FONT
        cell.fill = FRAME_HEADER_FILL
        cell.alignment = FRAME_HEADER_ALIGNMENT
        cell.border = FRAME_HEADER_BORDER
    number_formats = [_number_format(column) for column in columns]
    widths = [max(10, len(column) + 2) for column in columns]
    row_count = 0
    for row_index, row in enumerate(selected.iter_rows(named=False), start=5):
        row_count += 1
        alternate_row = row_index % 2 == 0
        for column_index, (column, value) in enumerate(
            zip(columns, row), start=1
        ):
            cell = ws.cell(row_index, column_index, _excel_value(value))
            cell.font = FRAME_BODY_FONT
            cell.border = FRAME_BODY_BORDER
            if column == "summary_description":
                cell.alignment = FRAME_SUMMARY_ALIGNMENT
                ws.row_dimensions[row_index].height = 54
            if isinstance(cell.value, (int, float, date, datetime)):
                cell.number_format = number_formats[column_index - 1]
            if alternate_row:
                cell.fill = FRAME_ALTERNATE_FILL
            if row_count <= 200:
                widths[column_index - 1] = min(
                    42,
                    max(
                        widths[column_index - 1],
                        len(str(value or "")) + 1,
                    ),
                )
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(columns))}{max(4, 4 + row_count)}"
    if row_count:
        table = Table(
            displayName=_table_name(name),
            ref=f"A4:{get_column_letter(len(columns))}{4 + row_count}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.sheet_view.showGridLines = False
    status_columns = [
        index for index, column in enumerate(columns, start=1)
        if column.lower() in {"status", "oos_grade", "liquidity_gate"}
    ]
    for column_index in status_columns:
        letter = get_column_letter(column_index)
        data_range = f"{letter}5:{letter}{max(5, 4 + row_count)}"
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'OR(ISNUMBER(SEARCH("PASS",{letter}5)),ISNUMBER(SEARCH("VALIDATED",{letter}5)),ISNUMBER(SEARCH("BUY",{letter}5)))'],
                fill=PatternFill("solid", fgColor=PALE_GREEN),
            ),
        )
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'OR(ISNUMBER(SEARCH("FAIL",{letter}5)),ISNUMBER(SEARCH("SELL",{letter}5)),ISNUMBER(SEARCH("BLOCK",{letter}5)))'],
                fill=PatternFill("solid", fgColor=PALE_RED),
            ),
        )
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(
                formula=[f'OR(ISNUMBER(SEARCH("WARN",{letter}5)),ISNUMBER(SEARCH("WATCH",{letter}5)),ISNUMBER(SEARCH("RESEARCH",{letter}5)))'],
                fill=PatternFill("solid", fgColor=PALE_AMBER),
            ),
        )
    _configure_print(ws, max(5, 4 + row_count), len(columns))
    return ws


def _control_sheet(
    workbook: Workbook,
    signals: pl.DataFrame,
    manifest: dict[str, object],
    demo_mode: bool,
) -> None:
    ws = workbook.active
    ws.title = "00 Control"
    _title_block(
        ws,
        "Bloomberg Distillate Relative-Value Technical System",
        "Windows-first XBBG + Polars workflow • train/evaluate/freeze or fast score-only • M1–M16 • D-4 liquidation",
        16,
    )
    ws.merge_cells("A3:P3")
    ws["A3"] = (
        "DEMO / SYNTHETIC PAPER DATA — no directional recommendation can be promoted."
        if demo_mode
        else "LIVE BLOOMBERG DATA — every signal remains subject to evidence, liquidity, and expiry gates."
    )
    ws["A3"].font = Font(bold=True, color="704B00")
    ws["A3"].fill = PatternFill("solid", fgColor=PALE_AMBER)
    ws["A3"].alignment = Alignment(wrap_text=True)
    status_counts = {
        str(row["status"]): int(row["count"])
        for row in signals.group_by("status")
        .agg(pl.len().alias("count"))
        .to_dicts()
    }
    model_id = str(manifest.get("model_id", "n/a"))
    model_parts = model_id.split("-")
    model_display = "-".join(model_parts[:2]) if len(model_parts) >= 2 else model_id[:16]
    kpis = [
        ("Structures", signals.height),
        ("Actionable", status_counts.get("BUY", 0) + status_counts.get("SELL", 0)),
        ("Watch", status_counts.get("WATCH", 0)),
        ("Blocked / no trade", sum(value for key, value in status_counts.items() if "BLOCK" in key or key == "NO TRADE")),
        ("Model", model_display),
        ("Workflow", manifest.get("workflow", "n/a")),
        ("As of", manifest.get("as_of", "n/a")),
        ("Validation", "PASS"),
    ]
    for index, (label, value) in enumerate(kpis):
        column = 1 + index * 2
        ws.merge_cells(start_row=5, start_column=column, end_row=5, end_column=column + 1)
        ws.merge_cells(start_row=6, start_column=column, end_row=6, end_column=column + 1)
        ws.cell(5, column, label)
        ws.cell(6, column, _excel_value(value))
        ws.cell(5, column).font = Font(size=9, bold=True, color=GRAY)
        ws.cell(6, column).font = Font(size=14, bold=True, color=NAVY)
        for row in (5, 6):
            ws.cell(row, column).fill = PatternFill("solid", fgColor=PALE_BLUE)
    headers = [
        "Status", "Trade code", "Family", "Current", "Buy ≤", "Sell ≥", "Fair value", "Unit",
        "Confidence", "Rel vol", "Pkg cap", "Kronos 1-bar", "Kronos status",
        "Last exit", "Evidence", "Model strategy",
    ]
    for index, value in enumerate(headers, start=1):
        ws.cell(9, index, value)
        ws.cell(9, index).font = Font(bold=True, color=WHITE)
        ws.cell(9, index).fill = PatternFill("solid", fgColor=BLUE)
    fields = [
        "status", "trade_code", "family", "display_current", "display_buy_entry",
        "display_sell_entry", "display_fair_value", "display_unit", "confidence", "relative_volume",
        "package_volume_capacity", "kronos_expected_move_1b", "kronos_status",
        "mandatory_last_exit_session", "oos_grade", "signal_strategy_id",
    ]
    for row_index, row in enumerate(signals.head(15).to_dicts(), start=10):
        for column_index, field in enumerate(fields, start=1):
            value = row.get(field)
            if field == "kronos_status":
                value = "OFF" if value == "DISABLED_OR_NOT_RUN" else "EXPERIMENTAL"
            elif field == "oos_grade" and str(value).startswith("DEMO ONLY"):
                value = "DEMO ONLY"
            cell = ws.cell(row_index, column_index, _excel_value(value))
            cell.border = Border(bottom=LINE)
            if isinstance(cell.value, (float, int, date, datetime)):
                cell.number_format = _number_format(field)
    ws.freeze_panes = "A9"
    ws.sheet_view.showGridLines = False
    for column in range(1, 17):
        ws.column_dimensions[get_column_letter(column)].width = 16 if column != 2 else 28
    _configure_print(ws, ws.max_row, 16)


def _checks_frame(
    signals: pl.DataFrame,
    scorecard: pl.DataFrame,
    quality: pl.DataFrame,
    coverage: pl.DataFrame,
    audit: pl.DataFrame,
    adaptive: pl.DataFrame,
    trades: pl.DataFrame,
    demo_mode: bool,
) -> pl.DataFrame:
    weight_columns = [column for column in adaptive.columns if column.startswith("adaptive_weight_")]
    weight_error = (
        float(adaptive.select((pl.sum_horizontal(*weight_columns) - 1.0).abs().max()).item())
        if weight_columns and not adaptive.is_empty() else 0.0
    )
    weight_max = (
        float(adaptive.select(pl.max_horizontal(*weight_columns).max()).item())
        if weight_columns and not adaptive.is_empty() else 0.0
    )
    checks = [
        ("Signals equal registered structures", signals.height, coverage.height, signals.height == coverage.height),
        ("All independent trials counted", scorecard.height, scorecard["trial_id"].n_unique() if not scorecard.is_empty() else 0, scorecard.is_empty() or scorecard.height == scorecard["trial_id"].n_unique()),
        ("Every structure observed", coverage.filter(pl.col("status") != "OBSERVED").height if not coverage.is_empty() else 0, 0, coverage.is_empty() or coverage.filter(pl.col("status") != "OBSERVED").is_empty()),
        ("Indicator audit has no missing package", audit.filter(pl.col("status") == "MISSING_PACKAGE").height if not audit.is_empty() else 0, 0, audit.is_empty() or audit.filter(pl.col("status") == "MISSING_PACKAGE").is_empty()),
        ("Adaptive weights sum to one", weight_error, 0.0, weight_error <= 1e-5),
        ("Adaptive expert cap", weight_max, 0.35, weight_max <= 0.350001),
        ("No blocking data-quality failure", quality.filter(pl.col("blocking") & (pl.col("status") == "FAIL")).height if not quality.is_empty() else 0, 0, quality.is_empty() or quality.filter(pl.col("blocking") & (pl.col("status") == "FAIL")).is_empty()),
        ("No demo BUY/SELL", signals.filter(pl.col("status").is_in(["BUY", "SELL"])).height if demo_mode else 0, 0, not demo_mode or signals.filter(pl.col("status").is_in(["BUY", "SELL"])).is_empty()),
        ("No trade exits after forced session", trades.filter(pl.col("exit_session") > pl.col("forced_exit_session")).height if not trades.is_empty() else 0, 0, trades.is_empty() or trades.filter(pl.col("exit_session") > pl.col("forced_exit_session")).is_empty()),
    ]
    return pl.DataFrame(
        [
            {
                "check": name,
                "actual": actual,
                "expected": expected,
                "status": "OK" if passed else "FAIL",
            }
            for name, actual, expected, passed in checks
        ]
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--project-root", type=Path, default=PROJECT_ROOT)
    parser.add_argument("--mode", choices=("live", "demo"), required=True)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    root = args.project_root.resolve()
    dist = root / "dist"
    output = args.output.resolve() if args.output else root / "Technical_Trading_System.xlsx"
    manifest_path = dist / "technical_run_manifest.json"
    scoring_manifest = dist / "technical_scoring_manifest.json"
    if scoring_manifest.is_file() and scoring_manifest.stat().st_mtime >= manifest_path.stat().st_mtime:
        manifest_path = scoring_manifest
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    demo_mode = args.mode == "demo"

    signals = _frame(dist / "technical_live_signals.csv")
    structure_summaries = _frame(dist / "technical_structure_summaries.csv")
    model_summary_path = dist / "technical_model_summary.json"
    model_summary = (
        json.loads(model_summary_path.read_text(encoding="utf-8"))
        if model_summary_path.is_file()
        else {}
    )
    scorecard = _frame(dist / "technical_strategy_scorecard.csv")
    strategies = _frame(dist / "technical_strategy_library.csv")
    trades = _frame(dist / "technical_backtest_trades.csv")
    indicators = _frame(dist / "technical_current_indicators.csv")
    spread_library = _frame(dist / "technical_spread_library.csv")
    spread_legs = _frame(dist / "technical_spread_legs.csv")
    expiry = _frame(dist / "technical_expiry_calendar.csv")
    quality = _frame(dist / "technical_data_quality.csv")
    folds = _frame(dist / "technical_fold_metrics.csv")
    parameters = _frame(dist / "technical_parameter_catalog.csv")
    history = _frame(dist / "technical_daily_spread_history.csv")
    adaptive = _frame(dist / "technical_adaptive_weight_history.csv")
    seasonality = _frame(dist / "technical_seasonality_profiles.csv")
    coverage = _frame(dist / "technical_structure_coverage.csv")
    audit = _frame(dist / "technical_indicator_library_audit.csv")

    workbook = Workbook()
    _control_sheet(workbook, signals, manifest, demo_mode)
    brief_preferred = list(TRADE_BRIEF_PREFERRED_COLUMNS)
    workbook_briefs = structure_summaries
    if not structure_summaries.is_empty():
        workbook_briefs = structure_summaries.select(
            *[
                column
                for column in brief_preferred
                if column in structure_summaries.columns
            ],
            *[
                column
                for column in structure_summaries.columns
                if column not in brief_preferred
            ],
        )
    _write_frame_sheet(
        workbook,
        "01 Signals",
        "Current Trade Briefs",
        "Current levels, targets, selected OOS evidence, and latest-30-session results.",
        workbook_briefs if not workbook_briefs.is_empty() else signals,
    )
    _write_frame_sheet(workbook, "02 Scorecard", "Walk-Forward Strategy Scorecard", "All preregistered model trials, including zero-trade combinations.", scorecard)
    overview = pl.DataFrame(
        model_summary.get("selected_strategy_distribution") or [], strict=False
    )
    if overview.is_empty() and not scorecard.is_empty():
        overview = (
            scorecard.group_by(["strategy_id", "strategy_name"])
        .agg(
            pl.len().alias("structures"),
            (pl.col("status") == "VALIDATED").sum().alias("validated"),
            pl.col("oos_trades").sum().alias("oos_trades"),
            pl.col("net_pnl_usd").sum().alias("oos_net_pnl_usd"),
            pl.col("daily_sharpe").median().alias("median_daily_sharpe"),
            pl.col("deflated_sharpe_probability").median().alias("median_dsr_probability"),
        )
        .sort("oos_net_pnl_usd", descending=True)
        )
    overview_ws = _write_frame_sheet(workbook, "03 Strategy Overview", "Strategy Overview", "Aggregated OOS evidence; never a substitute for per-structure validation.", overview)
    if not overview.is_empty():
        chart = BarChart()
        chart.title = "OOS net P&L by strategy"
        chart.y_axis.title = "USD"
        chart.add_data(Reference(overview_ws, min_col=6, min_row=4, max_row=4 + overview.height), titles_from_data=True)
        chart.set_categories(Reference(overview_ws, min_col=2, min_row=5, max_row=4 + overview.height))
        chart.height = 7
        chart.width = 14
        overview_ws.add_chart(chart, "J4")
    core_history = history.filter(pl.col("core_spread")) if "core_spread" in history.columns else history.head(2000)
    history_ws = _write_frame_sheet(workbook, "04 Core History", "Core Spread History", "Recent current-contract-scale diagnostics for core structures.", core_history, limit=5000)
    if not core_history.is_empty() and {"spread_close", "session_date"}.issubset(core_history.columns):
        close_column = core_history.columns.index("spread_close") + 1
        date_column = core_history.columns.index("session_date") + 1
        chart = LineChart()
        chart.title = "Core spread history"
        chart.add_data(Reference(history_ws, min_col=close_column, min_row=4, max_row=min(4 + core_history.height, 300)), titles_from_data=True)
        chart.set_categories(Reference(history_ws, min_col=date_column, min_row=5, max_row=min(4 + core_history.height, 300)))
        chart.height = 7
        chart.width = 15
        history_ws.add_chart(chart, "R4")
    _write_frame_sheet(workbook, "05 Trade Log", "Auditable Trade Ledger", "Latest 5,000 next-bar fills with costs, phases, and expiry exits.", trades.sort("exit_time", descending=True) if not trades.is_empty() else trades, limit=5000)
    _write_frame_sheet(workbook, "06 Indicator Snapshot", "Current Indicator Snapshot", "Native grouped Polars features on the current executable contract scale.", indicators)
    _write_frame_sheet(workbook, "07 Spread Library", "Spread Registry", "Normalized display formulas and executable package economics.", spread_library)
    _write_frame_sheet(workbook, "08 Spread Legs", "Spread Leg Registry", "Exact leg signs, ratios, delivery offsets, and native economics.", spread_legs)
    _write_frame_sheet(workbook, "09 Strategy Library", "Preregistered Strategy Library", "Fixed rules; dynamic selection uses past OOS evidence rather than rule invention.", strategies)
    _write_frame_sheet(workbook, "10 Expiry Calendar", "Expiry and Forced-Exit Calendar", "Earliest verified leg risk controls every package.", expiry)
    _write_frame_sheet(workbook, "11 Data Quality", "Data-Quality Gates", "Blocking and non-blocking source checks.", quality)
    _write_frame_sheet(workbook, "12 Fold Metrics", "Walk-Forward Fold Metrics", "Chronological OOS fold evidence.", folds)
    _write_frame_sheet(workbook, "13 Parameters", "Parameter Catalog", "Run, liquidity, indicator, cost, and validation settings.", parameters)
    notes = pl.DataFrame(
        {
            "topic": [
                "Mode",
                "Workflow",
                "Model",
                "Lockbox",
                "Kronos",
                "Runtime",
                "Model summary",
                "Disclaimer",
            ],
            "detail": [
                str(manifest.get("mode")),
                str(manifest.get("workflow")),
                str(manifest.get("model_id")),
                json.dumps(manifest.get("model_training_window") or {}),
                "Optional isolated leg-level OHLCV diagnostic; never action-eligible in this release.",
                "Pandas-free operating path using XBBG/Polars/Arrow/NumPy.",
                str(model_summary.get("description") or "Model summary unavailable."),
                "Research and decision support only; not an autonomous order router or investment recommendation.",
            ],
        }
    )
    _write_frame_sheet(workbook, "14 Sources & Notes", "Sources and Operating Notes", "Provenance, model state, and required caveats.", notes)
    checks = _checks_frame(signals, scorecard, quality, coverage, audit, adaptive, trades, demo_mode)
    _write_frame_sheet(workbook, "15 Checks", "Workbook & Model Integrity Checks", "Independent static tie-outs sourced from validated output files.", checks)
    _write_frame_sheet(workbook, "16 History Data", "History Data", "Bounded recent history for workbook exploration; full detail remains in Parquet/CSV.", history.sort("session_date", descending=True) if not history.is_empty() and "session_date" in history.columns else history, limit=10000)
    latest_adaptive = adaptive.sort("session_date").group_by("spread_id", maintain_order=True).tail(1) if not adaptive.is_empty() else adaptive
    _write_frame_sheet(workbook, "17 Adaptive Weights", "Frozen Adaptive Weights", "Latest causal expert weights by structure.", latest_adaptive)
    latest_seasonality = seasonality.sort("day_of_year").group_by("spread_id", maintain_order=True).tail(1) if not seasonality.is_empty() else seasonality
    _write_frame_sheet(workbook, "18 Seasonality", "Prior-Year Seasonality", "Latest prior-year-only expected moves and support.", latest_seasonality)
    _write_frame_sheet(workbook, "19 Structure Coverage", "Structure Coverage", "Every registered structure remains visible and fails closed when incomplete.", coverage)
    _write_frame_sheet(workbook, "20 Indicator Audit", "Indicator Library Audit", "Native Polars versus lightweight Talis; Pandas-dependent comparisons are optional.", audit)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{output.stem}.", suffix=".xlsx", dir=str(output.parent)
    )
    os.close(descriptor)
    temp = Path(temp_name)
    try:
        workbook.save(temp)
        os.replace(temp, output)
    finally:
        if temp.exists():
            temp.unlink()
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
